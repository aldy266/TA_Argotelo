import io
import re
from pathlib import Path
from uuid import uuid4

from flask import Blueprint, request, jsonify, session, send_file
from datetime import datetime, timedelta, date, time
from functools import wraps
from sqlalchemy import func, text
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from model import db, Staff, Shift, StaffSchedule, Attendance, LeaveRequest, User
from utils.roles import expand_role_names, role_form_code, role_group, role_label, role_sort_key

staff_bp = Blueprint("staff_api", __name__, url_prefix="/api/staff")

STAFF_IMPORT_HEADERS = [
    "employee_code",
    "full_name",
    "department",
    "position",
    "phone",
    "email",
    "joined_at",
    "status",
]
STAFF_IMPORT_REQUIRED = ["employee_code", "full_name", "department"]
ALLOWED_STAFF_STATUS = {"ACTIVE", "INACTIVE"}
MAX_IMPORT_SIZE = 5 * 1024 * 1024
MAX_LEAVE_DOCUMENT_SIZE = 5 * 1024 * 1024
ALLOWED_LEAVE_DOCUMENT_EXTENSIONS = {"pdf", "jpg", "jpeg", "png", "webp", "doc", "docx"}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
BASE_DIR = Path(__file__).resolve().parents[1]
SAMPLE_DATA_DIR = BASE_DIR / "sample_data"
DUMMY_STAFF_FILE = SAMPLE_DATA_DIR / "data_dummy_staff.xlsx"
LEAVE_DOCUMENT_DIR = BASE_DIR / "static" / "uploads" / "leave_documents"

# ====================================================
# AUTHORIZATION HELPERS
# ====================================================

def require_auth():
    """Validate user is authenticated"""
    if "user_id" not in session:
        return None
    return session.get("user_id")

def require_role(*roles):
    """Validate user has required role"""
    user_id = require_auth()
    if not user_id:
        return None
    
    user = User.query.get(user_id)
    if not user:
        return None
    
    allowed_roles = expand_role_names(roles)
    user_role = user.role.role_name.upper()

    if user_role not in allowed_roles:
        return None
    
    return user

def check_authorization(*required_roles):
    """Decorator to check authorization"""
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            user = require_role(*required_roles)
            if not user:
                status = 401 if "user_id" not in session else 403
                return jsonify({
                    "success": False,
                    "message": "Belum login" if status == 401 else "Akses ditolak"
                }), status
            return f(user, *args, **kwargs)
        return wrapper
    return decorator

def waktu_wib():
    """Indonesia timezone"""
    return datetime.utcnow() + timedelta(hours=7)


def normalize_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value):
    return normalize_cell(value).lower()


def parse_joined_at(value):
    if value in (None, ""):
        return datetime.combine(waktu_wib().date(), time.min)

    if isinstance(value, datetime):
        return datetime.combine(value.date(), time.min)

    if isinstance(value, date):
        return datetime.combine(value, time.min)

    value = normalize_cell(value)
    try:
        parsed = datetime.strptime(value, "%Y-%m-%d")
        return datetime.combine(parsed.date(), time.min)
    except ValueError as exc:
        raise ValueError("joined_at harus berformat YYYY-MM-DD") from exc


def parse_shift_time(value, field_label):
    value = normalize_cell(value)
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"{field_label} harus berformat HH:MM")


def get_today_shift_counts():
    today = waktu_wib().date()
    rows = db.session.query(
        StaffSchedule.shift_id,
        func.count(StaffSchedule.id)
    ).join(
        Staff, StaffSchedule.staff_id == Staff.id
    ).filter(
        StaffSchedule.schedule_date == today,
        Staff.status == "ACTIVE"
    ).group_by(
        StaffSchedule.shift_id
    ).all()

    return {shift_id: count for shift_id, count in rows}


def serialize_shift(shift, today_staff_count=None):
    return {
        "id": shift.id,
        "shift_name": shift.shift_name,
        "start_time": shift.start_time.isoformat(),
        "end_time": shift.end_time.isoformat(),
        "tolerance_minutes": shift.tolerance_minutes,
        "today_staff_count": int(today_staff_count or 0)
    }


def sync_attendance_with_schedule(schedule, attendance=None):
    attendance = attendance or Attendance.query.filter_by(schedule_id=schedule.id).first()
    if not attendance:
        attendance = Attendance(
            staff_id=schedule.staff_id,
            schedule_id=schedule.id,
            attendance_date=schedule.schedule_date,
            status="NOT_CHECKED_IN"
        )
        db.session.add(attendance)
        return attendance

    attendance.staff_id = schedule.staff_id
    attendance.attendance_date = schedule.schedule_date

    if attendance.status in ["LEAVE", "SICK", "ABSENT"]:
        return attendance

    if not attendance.clock_in:
        attendance.status = "NOT_CHECKED_IN"
        attendance.late_minutes = 0
        attendance.work_minutes = attendance.work_minutes or 0
        return attendance

    shift_start = datetime.combine(schedule.schedule_date, schedule.shift.start_time)
    tolerance_limit = shift_start + timedelta(minutes=schedule.shift.tolerance_minutes or 0)
    late_minutes = max(int((attendance.clock_in - tolerance_limit).total_seconds() // 60), 0)
    attendance.late_minutes = late_minutes

    if attendance.clock_out:
        attendance.work_minutes = calculate_work_minutes(
            attendance.clock_in,
            attendance.clock_out,
            attendance.work_minutes
        )
        attendance.status = "COMPLETED"
    else:
        attendance.status = "LATE" if late_minutes else "PRESENT"

    return attendance


def calculate_work_minutes(clock_in, clock_out, stored_minutes=0):
    if stored_minutes and stored_minutes > 0:
        return int(stored_minutes)

    if not clock_in or not clock_out:
        return 0

    return max(int((clock_out - clock_in).total_seconds() // 60), 0)


def schedule_datetime_range(schedule):
    start_at = datetime.combine(schedule.schedule_date, schedule.shift.start_time)
    end_at = datetime.combine(schedule.schedule_date, schedule.shift.end_time)
    if end_at <= start_at:
        end_at += timedelta(days=1)
    return start_at, end_at


def leave_attendance_status(leave_type):
    return "SICK" if leave_type == "SICK" else "LEAVE"


def apply_leave_request_to_attendance(leave_req):
    attendance_status = leave_attendance_status(leave_req.leave_type)
    current_date = leave_req.start_date
    applied_count = 0

    while current_date <= leave_req.end_date:
        schedules = StaffSchedule.query.filter_by(
            staff_id=leave_req.staff_id,
            schedule_date=current_date
        ).all()

        for schedule in schedules:
            attendance = Attendance.query.filter_by(schedule_id=schedule.id).first()
            if not attendance:
                attendance = Attendance(
                    staff_id=leave_req.staff_id,
                    schedule_id=schedule.id,
                    attendance_date=current_date
                )
                db.session.add(attendance)

            attendance.status = attendance_status
            attendance.clock_in = None
            attendance.clock_out = None
            attendance.late_minutes = 0
            attendance.work_minutes = 0
            applied_count += 1

        current_date += timedelta(days=1)

    return applied_count


def get_schedule_attendance(schedule):
    attendance = Attendance.query.filter_by(schedule_id=schedule.id).first()
    if attendance:
        return attendance

    attendance = Attendance(
        staff_id=schedule.staff_id,
        schedule_id=schedule.id,
        attendance_date=schedule.schedule_date,
        status="NOT_CHECKED_IN"
    )
    db.session.add(attendance)
    return attendance


def clock_in_schedule(schedule, now):
    shift_start, shift_end = schedule_datetime_range(schedule)

    if now < shift_start:
        return None, (
            f"Absen masuk belum dibuka. Jam kerja dimulai pukul {schedule.shift.start_time.strftime('%H:%M')}",
            403
        )

    if now > shift_end:
        return None, (
            "Absen masuk sudah ditutup karena jam kerja telah selesai",
            403
        )

    attendance = get_schedule_attendance(schedule)

    if attendance.status in ["LEAVE", "SICK", "ABSENT"]:
        return None, ("Status kehadiran tidak dapat clock in", 409)

    if attendance.clock_in:
        return None, ("Staff sudah clock in", 409)

    tolerance_limit = shift_start + timedelta(
        minutes=schedule.shift.tolerance_minutes or 0
    )
    late_minutes = max(int((now - tolerance_limit).total_seconds() // 60), 0)

    attendance.clock_in = now
    attendance.status = "LATE" if late_minutes else "PRESENT"
    attendance.late_minutes = late_minutes

    return attendance, None


def clock_out_schedule(schedule, now):
    attendance = Attendance.query.filter_by(schedule_id=schedule.id).first()
    _shift_start, shift_end = schedule_datetime_range(schedule)
    if not attendance or not attendance.clock_in:
        return None, ("Staff belum clock in", 409)

    if attendance.clock_out:
        return None, ("Staff sudah clock out", 409)

    if attendance.status in ["LEAVE", "SICK", "ABSENT"]:
        return None, ("Status kehadiran tidak dapat clock out", 409)

    if now < shift_end:
        return None, (
            f"Absen pulang belum dibuka. Jam kerja selesai pukul {schedule.shift.end_time.strftime('%H:%M')}",
            403
        )

    attendance.clock_out = now
    attendance.work_minutes = int((now - attendance.clock_in).total_seconds() // 60)
    attendance.status = "COMPLETED"

    return attendance, None


def format_staff_import_sheet(workbook, sheet_name):
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(STAFF_IMPORT_HEADERS)

    header_fill = PatternFill("solid", fgColor="5A3718")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"

    widths = {
        "A": 18,
        "B": 24,
        "C": 18,
        "D": 20,
        "E": 18,
        "F": 28,
        "G": 14,
        "H": 14,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    return sheet


def build_staff_template_workbook():
    workbook = Workbook()
    sheet = format_staff_import_sheet(workbook, "Staff Import")
    sheet.append([
        "STF001",
        "Nama Staff",
        "TIM TOKO",
        "Kasir",
        "081234567890",
        "staff@example.com",
        "2026-01-15",
        "ACTIVE",
    ])

    guide = workbook.create_sheet("Petunjuk")
    guide.append(["Kolom", "Keterangan"])
    guide["A1"].font = Font(bold=True, color="FFFFFF")
    guide["B1"].font = Font(bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor="5A3718")
    guide["B1"].fill = PatternFill("solid", fgColor="5A3718")
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 70
    guide.freeze_panes = "A2"
    guide_rows = [
        ("employee_code", "Kode Staff unik dan wajib."),
        ("full_name", "Nama lengkap Staff dan wajib."),
        ("department", "Divisi Staff dan wajib."),
        ("position", "Jabatan Staff."),
        ("phone", "Nomor telepon."),
        ("email", "Email Staff."),
        ("joined_at", "Format YYYY-MM-DD."),
        ("status", "ACTIVE atau INACTIVE. Jika kosong akan menjadi ACTIVE."),
    ]
    for row in guide_rows:
        guide.append(row)

    return workbook


def build_dummy_staff_workbook():
    workbook = Workbook()
    sheet = format_staff_import_sheet(workbook, "Staff Dummy Data")
    dummy_rows = [
        ["STF001", "Ahmad Ridwan", "TIM TOKO", "Kasir", "081234567801", "ahmad.ridwan@example.com", "2025-01-10", "ACTIVE"],
        ["STF002", "Siti Aminah", "TIM TOKO", "Kasir", "081234567802", "siti.aminah@example.com", "2025-01-15", "ACTIVE"],
        ["STF003", "Budi Santoso", "OPERASIONAL", "Supervisor Toko", "081234567803", "budi.santoso@example.com", "2025-02-01", "ACTIVE"],
        ["STF004", "Nadia Lestari", "FINANCE", "Staff Finance", "081234567804", "nadia.lestari@example.com", "2025-02-10", "ACTIVE"],
        ["STF005", "Rafi Hidayat", "TIM TOKO", "Kasir", "081234567805", "rafi.hidayat@example.com", "2025-02-17", "ACTIVE"],
        ["STF006", "Rina Kurnia", "HRD", "HR Staff", "081234567806", "rina.kurnia@example.com", "2025-03-01", "ACTIVE"],
        ["STF007", "Fajar Nugroho", "OPERASIONAL", "Staff Gudang", "081234567807", "fajar.nugroho@example.com", "2025-03-12", "ACTIVE"],
        ["STF008", "Nabila Rahma", "TIM TOKO", "Kasir", "081234567808", "nabila.rahma@example.com", "2025-04-01", "ACTIVE"],
        ["STF009", "Reza Maulana", "FINANCE", "Staff Finance", "081234567809", "reza.maulana@example.com", "2025-04-15", "ACTIVE"],
        ["STF010", "Intan Permata", "OPERASIONAL", "Staff Gudang", "081234567810", "intan.permata@example.com", "2025-05-02", "ACTIVE"],
    ]

    for row in dummy_rows:
        sheet.append(row)

    for row in sheet.iter_rows(min_row=2, max_col=len(STAFF_IMPORT_HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    return workbook


def workbook_response(workbook, filename):
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def validate_import_file(file_storage):
    if not file_storage or not file_storage.filename:
        return "File Excel belum dipilih."

    if not file_storage.filename.lower().endswith(".xlsx"):
        return "Format file harus .xlsx."

    content_length = request.content_length or 0
    if content_length > MAX_IMPORT_SIZE:
        return "Ukuran file maksimal 5 MB."

    current_pos = file_storage.stream.tell()
    file_storage.stream.seek(0, io.SEEK_END)
    size = file_storage.stream.tell()
    file_storage.stream.seek(current_pos)

    if size > MAX_IMPORT_SIZE:
        return "Ukuran file maksimal 5 MB."

    return None


def validate_leave_document(file_storage):
    if not file_storage or not file_storage.filename:
        return None

    filename = secure_filename(file_storage.filename)
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if extension not in ALLOWED_LEAVE_DOCUMENT_EXTENSIONS:
        return "Format surat izin harus PDF, gambar, DOC, atau DOCX."

    current_pos = file_storage.stream.tell()
    file_storage.stream.seek(0, io.SEEK_END)
    size = file_storage.stream.tell()
    file_storage.stream.seek(current_pos)

    if size > MAX_LEAVE_DOCUMENT_SIZE:
        return "Ukuran surat izin maksimal 5 MB."

    return None


def save_leave_document(file_storage):
    if not file_storage or not file_storage.filename:
        return None

    validation_error = validate_leave_document(file_storage)
    if validation_error:
        raise ValueError(validation_error)

    LEAVE_DOCUMENT_DIR.mkdir(parents=True, exist_ok=True)
    filename = secure_filename(file_storage.filename)
    extension = filename.rsplit(".", 1)[-1].lower()
    saved_name = f"{waktu_wib().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:10]}.{extension}"
    save_path = LEAVE_DOCUMENT_DIR / saved_name
    file_storage.save(save_path)
    return f"/static/uploads/leave_documents/{saved_name}"


def serialize_leave_request(leave_req):
    return {
        "id": leave_req.id,
        "staff_id": leave_req.staff_id,
        "staff_name": leave_req.staff.full_name,
        "leave_type": leave_req.leave_type,
        "start_date": leave_req.start_date.isoformat(),
        "end_date": leave_req.end_date.isoformat(),
        "reason": leave_req.reason,
        "document_url": leave_req.document_url,
        "status": leave_req.status,
        "created_at": leave_req.created_at.isoformat() if leave_req.created_at else None,
        "reviewed_at": leave_req.reviewed_at.isoformat() if leave_req.reviewed_at else None,
        "reviewer": leave_req.reviewer.fullname if leave_req.reviewer else None
    }


def row_is_empty(row_values):
    return all(value in (None, "") or normalize_cell(value) == "" for value in row_values)

# ====================================================
# STAFF IMPORT EXCEL
# ====================================================

@staff_bp.route("/import-template", methods=["GET"])
@check_authorization("FINANCE")
def download_import_template(user):
    workbook = build_staff_template_workbook()
    return workbook_response(workbook, "template_import_staff.xlsx")


@staff_bp.route("/dummy-excel", methods=["GET"])
@check_authorization("FINANCE")
def download_dummy_staff_excel(user):
    if DUMMY_STAFF_FILE.exists():
        return send_file(
            DUMMY_STAFF_FILE,
            as_attachment=True,
            download_name="data_dummy_staff.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    workbook = build_dummy_staff_workbook()
    return workbook_response(workbook, "data_dummy_staff.xlsx")


@staff_bp.route("/import-excel", methods=["POST"])
@check_authorization("FINANCE")
def import_staff_excel(user):
    file_storage = request.files.get("file")
    validation_error = validate_import_file(file_storage)

    if validation_error:
        return jsonify({
            "success": False,
            "message": validation_error
        }), 400

    try:
        file_storage.stream.seek(0)
        workbook = load_workbook(file_storage.stream, data_only=True)
    except Exception:
        return jsonify({
            "success": False,
            "message": "File Excel tidak dapat dibaca"
        }), 400

    sheet = workbook.active

    if sheet.max_row < 1:
        return jsonify({
            "success": False,
            "message": "File Excel tidak memiliki header"
        }), 400

    raw_headers = [cell.value for cell in sheet[1]]
    headers = [normalize_header(value) for value in raw_headers]
    header_map = {
        header: index
        for index, header in enumerate(headers)
        if header
    }

    missing_columns = [
        column
        for column in STAFF_IMPORT_REQUIRED
        if column not in header_map
    ]

    if missing_columns:
        return jsonify({
            "success": False,
            "message": "Kolom wajib tidak lengkap",
            "missing_columns": missing_columns
        }), 400

    total_rows = 0
    imported = 0
    skipped = 0
    failed = 0
    errors = []
    seen_codes = set()
    staff_to_insert = []

    existing_codes = {
        staff.employee_code
        for staff in Staff.query.with_entities(Staff.employee_code).all()
    }

    for excel_row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2
    ):
        row_values = list(row)
        if row_is_empty(row_values):
            continue

        total_rows += 1

        def get_value(column):
            index = header_map.get(column)
            if index is None or index >= len(row_values):
                return ""
            return row_values[index]

        employee_code = normalize_cell(get_value("employee_code"))
        full_name = normalize_cell(get_value("full_name"))
        department = normalize_cell(get_value("department"))
        position = normalize_cell(get_value("position")) or "Staff"
        phone = normalize_cell(get_value("phone"))
        email = normalize_cell(get_value("email"))
        raw_joined_at = get_value("joined_at")
        status = normalize_cell(get_value("status")).upper() or "ACTIVE"

        row_errors = []

        if not employee_code:
            row_errors.append("employee_code wajib diisi")

        if not full_name:
            row_errors.append("full_name wajib diisi")

        if not department:
            row_errors.append("department wajib diisi")

        if employee_code and employee_code in seen_codes:
            row_errors.append("employee_code duplicate dalam file Excel")

        if email and not EMAIL_RE.match(email):
            row_errors.append("Email tidak valid")

        if status not in ALLOWED_STAFF_STATUS:
            row_errors.append("status harus ACTIVE atau INACTIVE")

        try:
            joined_at = parse_joined_at(raw_joined_at)
        except ValueError as error:
            joined_at = None
            row_errors.append(str(error))

        if row_errors:
            failed += 1
            errors.append({
                "row": excel_row_number,
                "employee_code": employee_code,
                "message": "; ".join(row_errors)
            })
            if employee_code:
                seen_codes.add(employee_code)
            continue

        seen_codes.add(employee_code)

        if employee_code in existing_codes:
            skipped += 1
            continue

        staff_to_insert.append(Staff(
            employee_code=employee_code,
            full_name=full_name,
            department=department,
            position=position,
            phone=phone or None,
            email=email or None,
            joined_at=joined_at,
            status=status
        ))
        existing_codes.add(employee_code)
        imported += 1

    try:
        if staff_to_insert:
            db.session.add_all(staff_to_insert)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": "Import gagal disimpan ke database"
        }), 500

    return jsonify({
        "success": True,
        "message": "Import data Staff selesai",
        "data": {
            "total_rows": total_rows,
            "imported": imported,
            "skipped": skipped,
            "failed": failed,
            "errors": errors
        }
    }), 200


# ====================================================
# STAFF MANAGEMENT
# ====================================================

@staff_bp.route("", methods=["GET"])
@check_authorization("OWNER", "FINANCE", "HRD")
def get_staff(user):
    """Get all active staff"""
    try:
        include_inactive = request.args.get("include_inactive") == "1"
        query = Staff.query

        if not include_inactive:
            query = query.filter_by(status="ACTIVE")

        staff = query.order_by(Staff.full_name.asc()).all()
        
        data = [{
            "id": s.id,
            "employee_code": s.employee_code,
            "full_name": s.full_name,
            "department": s.department,
            "position": s.position,
            "phone": s.phone,
            "email": s.email,
            "joined_at": s.joined_at.isoformat() if s.joined_at else None,
            "status": s.status,
            "created_at": s.created_at.isoformat() if s.created_at else None
        } for s in staff]
        
        return jsonify({
            "success": True,
            "data": data
        }), 200
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@staff_bp.route("", methods=["POST"])
@check_authorization("FINANCE")
def create_staff(user):
    """Create new staff"""
    try:
        data = request.get_json()
        
        # Validasi
        if not data.get("full_name") or not data.get("employee_code"):
            return jsonify({
                "success": False,
                "message": "Nama lengkap dan kode karyawan wajib diisi"
            }), 400
        
        if not data.get("department") or not data.get("position"):
            return jsonify({
                "success": False,
                "message": "Divisi dan jabatan wajib diisi"
            }), 400
        
        # Cek duplikasi employee_code
        existing = Staff.query.filter_by(employee_code=data["employee_code"]).first()
        if existing:
            return jsonify({
                "success": False,
                "message": "Kode karyawan sudah terdaftar"
            }), 409
        
        # Cek email unique jika diberikan
        if data.get("email"):
            existing_email = Staff.query.filter_by(email=data["email"]).first()
            if existing_email:
                return jsonify({
                    "success": False,
                    "message": "Email sudah terdaftar"
                }), 409
        
        joined_at = datetime.fromisoformat(data.get("joined_at", waktu_wib().isoformat()))
        
        staff = Staff(
            full_name=data["full_name"],
            employee_code=data["employee_code"],
            department=data["department"],
            position=data["position"],
            phone=data.get("phone"),
            email=data.get("email"),
            joined_at=joined_at,
            status="ACTIVE"
        )
        
        db.session.add(staff)
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"{staff.full_name} berhasil ditambahkan",
            "data": {
                "id": staff.id,
                "employee_code": staff.employee_code,
                "full_name": staff.full_name
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@staff_bp.route("/<int:staff_id>", methods=["PUT"])
@check_authorization("OWNER", "FINANCE")
def update_staff(user, staff_id):
    """Update staff"""
    try:
        staff = Staff.query.get(staff_id)
        if not staff:
            return jsonify({
                "success": False,
                "message": "Staff tidak ditemukan"
            }), 404
        
        data = request.get_json()
        
        # Update fields yang dapat diubah
        if "full_name" in data:
            staff.full_name = data["full_name"]
        if "department" in data:
            staff.department = data["department"]
        if "position" in data:
            staff.position = data["position"]
        if "phone" in data:
            staff.phone = data["phone"]
        if "email" in data:
            staff.email = data["email"]
        if "joined_at" in data:
            staff.joined_at = datetime.fromisoformat(data["joined_at"])
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "Staff berhasil diupdate",
            "data": {
                "id": staff.id,
                "full_name": staff.full_name
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@staff_bp.route("/<int:staff_id>/deactivate", methods=["PATCH"])
@check_authorization("OWNER", "FINANCE")
def deactivate_staff(user, staff_id):
    """Deactivate staff (soft delete)"""
    try:
        staff = Staff.query.get(staff_id)
        if not staff:
            return jsonify({
                "success": False,
                "message": "Staff tidak ditemukan"
            }), 404
        
        staff.status = "INACTIVE"
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"{staff.full_name} berhasil dinonaktifkan"
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

# ====================================================
# SHIFT MANAGEMENT
# ====================================================

@staff_bp.route("/shift", methods=["GET"])
@check_authorization("OWNER", "FINANCE", "HRD")
def get_shifts(user):
    """Get all active shifts"""
    try:
        shifts = Shift.query.filter_by(status="ACTIVE").order_by(Shift.start_time.asc()).all()
        today_counts = get_today_shift_counts()
        
        return jsonify({
            "success": True,
            "data": [
                serialize_shift(shift, today_counts.get(shift.id, 0))
                for shift in shifts
            ]
        }), 200
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/shift", methods=["POST"])
@check_authorization("OWNER")
def create_shift(user):
    """Create a new active shift."""
    try:
        data = request.get_json(silent=True) or {}

        shift_name = normalize_cell(data.get("shift_name"))
        if not shift_name:
            return jsonify({
                "success": False,
                "message": "Nama shift wajib diisi"
            }), 400

        start_time = parse_shift_time(data.get("start_time"), "Jam mulai")
        end_time = parse_shift_time(data.get("end_time"), "Jam selesai")
        if start_time == end_time:
            return jsonify({
                "success": False,
                "message": "Jam mulai dan jam selesai tidak boleh sama"
            }), 400

        try:
            tolerance_minutes = int(data.get("tolerance_minutes") or 0)
        except (TypeError, ValueError):
            return jsonify({
                "success": False,
                "message": "Toleransi harus berupa angka"
            }), 400

        if tolerance_minutes < 0:
            return jsonify({
                "success": False,
                "message": "Toleransi tidak boleh kurang dari 0"
            }), 400

        shift = Shift(
            shift_name=shift_name,
            start_time=start_time,
            end_time=end_time,
            tolerance_minutes=tolerance_minutes,
            status="ACTIVE"
        )
        db.session.add(shift)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Shift berhasil ditambahkan",
            "data": serialize_shift(shift, 0)
        }), 201
    except ValueError as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/shift/<int:shift_id>", methods=["PATCH"])
@check_authorization("OWNER")
def update_shift(user, shift_id):
    """Update shift name and working hours"""
    try:
        shift = Shift.query.get(shift_id)
        if not shift or shift.status != "ACTIVE":
            return jsonify({
                "success": False,
                "message": "Shift tidak ditemukan"
            }), 404

        data = request.get_json(silent=True) or {}

        if "shift_name" in data:
            shift_name = normalize_cell(data.get("shift_name"))
            if not shift_name:
                return jsonify({
                    "success": False,
                    "message": "Nama shift wajib diisi"
                }), 400
            shift.shift_name = shift_name

        if "start_time" in data:
            shift.start_time = parse_shift_time(data.get("start_time"), "Jam mulai")

        if "end_time" in data:
            shift.end_time = parse_shift_time(data.get("end_time"), "Jam selesai")

        if shift.start_time == shift.end_time:
            return jsonify({
                "success": False,
                "message": "Jam mulai dan jam selesai tidak boleh sama"
            }), 400

        if "tolerance_minutes" in data:
            try:
                tolerance_minutes = int(data.get("tolerance_minutes") or 0)
            except (TypeError, ValueError):
                return jsonify({
                    "success": False,
                    "message": "Toleransi harus berupa angka"
                }), 400

            if tolerance_minutes < 0:
                return jsonify({
                    "success": False,
                    "message": "Toleransi tidak boleh kurang dari 0"
                }), 400

            shift.tolerance_minutes = tolerance_minutes

        today = waktu_wib().date()
        today_schedules = StaffSchedule.query.filter_by(
            shift_id=shift.id,
            schedule_date=today
        ).all()
        for schedule in today_schedules:
            schedule.shift = shift
            attendance = Attendance.query.filter_by(schedule_id=schedule.id).first()
            if attendance:
                sync_attendance_with_schedule(schedule, attendance)

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Shift berhasil diperbarui",
            "data": serialize_shift(
                shift,
                get_today_shift_counts().get(shift.id, 0)
            )
        }), 200
    except ValueError as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

# ====================================================
# SCHEDULE MANAGEMENT
# ====================================================

@staff_bp.route("/schedule", methods=["POST"])
@check_authorization("OWNER", "FINANCE")
def create_schedule(user):
    """Create staff schedule"""
    try:
        data = request.get_json(silent=True) or {}
        
        if not data.get("staff_id") or not data.get("shift_id"):
            return jsonify({
                "success": False,
                "message": "Staff dan Shift wajib dipilih"
            }), 400

        if not data.get("schedule_date"):
            return jsonify({
                "success": False,
                "message": "Tanggal mulai wajib diisi"
            }), 400

        staff_id_value = str(data.get("staff_id")).strip()
        shift_id = int(data.get("shift_id"))
        schedule_date = datetime.fromisoformat(data.get("schedule_date")).date()
        end_date_value = data.get("end_date") or data.get("schedule_end_date")
        end_date = (
            datetime.fromisoformat(end_date_value).date()
            if end_date_value
            else schedule_date
        )

        if end_date < schedule_date:
            return jsonify({
                "success": False,
                "message": "Tanggal selesai tidak boleh sebelum tanggal mulai"
            }), 400

        total_days = (end_date - schedule_date).days + 1
        if total_days > 366:
            return jsonify({
                "success": False,
                "message": "Rentang jadwal maksimal 1 tahun"
            }), 400

        if staff_id_value.lower() == "all":
            staff_ids = [
                staff.id
                for staff in Staff.query.filter_by(status="ACTIVE").order_by(Staff.full_name.asc()).all()
            ]
            if not staff_ids:
                return jsonify({
                    "success": False,
                    "message": "Belum ada Staff aktif"
                }), 404
        else:
            staff_id = int(staff_id_value)
            staff = Staff.query.get(staff_id)
            if not staff or staff.status != "ACTIVE":
                return jsonify({
                    "success": False,
                    "message": "Staff tidak ditemukan atau tidak aktif"
                }), 404
            staff_ids = [staff_id]

        shift = Shift.query.get(shift_id)
        if not shift or shift.status != "ACTIVE":
            return jsonify({
                "success": False,
                "message": "Shift tidak ditemukan"
            }), 404

        created = 0
        skipped = 0
        first_schedule = None

        for staff_id in staff_ids:
            for offset in range(total_days):
                current_date = schedule_date + timedelta(days=offset)
                existing_schedule = StaffSchedule.query.filter_by(
                    staff_id=staff_id,
                    schedule_date=current_date
                ).first()

                if existing_schedule:
                    if len(staff_ids) == 1 and total_days == 1:
                        return jsonify({
                            "success": False,
                            "message": "Staff sudah memiliki jadwal pada tanggal tersebut."
                        }), 409
                    skipped += 1
                    continue

                schedule = StaffSchedule(
                    staff_id=staff_id,
                    shift_id=shift_id,
                    schedule_date=current_date
                )
                db.session.add(schedule)
                db.session.flush()

                db.session.add(Attendance(
                    staff_id=staff_id,
                    schedule_id=schedule.id,
                    attendance_date=current_date,
                    status="NOT_CHECKED_IN"
                ))

                first_schedule = first_schedule or schedule
                created += 1

        if created == 0:
            return jsonify({
                "success": False,
                "message": "Tidak ada jadwal yang dibuat. Semua Staff sudah memiliki jadwal pada tanggal tersebut."
            }), 409

        db.session.commit()

        if len(staff_ids) == 1 and total_days == 1:
            message = "Jadwal berhasil ditambahkan"
        else:
            message = f"{created} jadwal berhasil dibuat"
            if skipped:
                message += f", {skipped} jadwal dilewati karena sudah ada"

        return jsonify({
            "success": True,
            "message": message,
            "data": {
                "id": first_schedule.id if first_schedule else None,
                "schedule_date": schedule_date.isoformat(),
                "end_date": end_date.isoformat(),
                "created": created,
                "skipped": skipped,
                "staff_count": len(staff_ids),
                "total_days": total_days
            }
        }), 201
    except ValueError:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": "Data jadwal tidak valid"
        }), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/schedule/<int:schedule_id>", methods=["PATCH", "PUT"])
@check_authorization("OWNER", "FINANCE")
def update_schedule(user, schedule_id):
    """Update a staff schedule"""
    try:
        schedule = StaffSchedule.query.get(schedule_id)
        if not schedule:
            return jsonify({
                "success": False,
                "message": "Jadwal tidak ditemukan"
            }), 404

        data = request.get_json(silent=True) or {}

        staff_id = int(data.get("staff_id") or schedule.staff_id)
        shift_id = int(data.get("shift_id") or schedule.shift_id)
        schedule_date_value = data.get("schedule_date")
        schedule_date = (
            datetime.fromisoformat(schedule_date_value).date()
            if schedule_date_value
            else schedule.schedule_date
        )

        staff = Staff.query.get(staff_id)
        if not staff or staff.status != "ACTIVE":
            return jsonify({
                "success": False,
                "message": "Staff tidak ditemukan atau tidak aktif"
            }), 404

        shift = Shift.query.get(shift_id)
        if not shift or shift.status != "ACTIVE":
            return jsonify({
                "success": False,
                "message": "Shift tidak ditemukan"
            }), 404

        existing_schedule = StaffSchedule.query.filter(
            StaffSchedule.staff_id == staff_id,
            StaffSchedule.schedule_date == schedule_date,
            StaffSchedule.id != schedule.id
        ).first()
        if existing_schedule:
            return jsonify({
                "success": False,
                "message": "Staff sudah memiliki jadwal pada tanggal tersebut."
            }), 409

        schedule.staff_id = staff_id
        schedule.shift_id = shift_id
        schedule.shift = shift
        schedule.schedule_date = schedule_date

        attendance = Attendance.query.filter_by(schedule_id=schedule.id).first()
        sync_attendance_with_schedule(schedule, attendance)

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Jadwal berhasil diperbarui",
            "data": {
                "id": schedule.id,
                "staff_id": schedule.staff_id,
                "shift_id": schedule.shift_id,
                "schedule_date": schedule.schedule_date.isoformat()
            }
        }), 200
    except ValueError:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": "Data jadwal tidak valid"
        }), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/schedule/<int:schedule_id>", methods=["DELETE"])
@check_authorization("OWNER", "FINANCE")
def delete_schedule(user, schedule_id):
    """Delete a staff schedule and its attendance record"""
    try:
        schedule = StaffSchedule.query.get(schedule_id)
        if not schedule:
            return jsonify({
                "success": False,
                "message": "Jadwal tidak ditemukan"
            }), 404

        db.session.delete(schedule)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Jadwal berhasil dihapus"
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

# ====================================================
# ATTENDANCE - GET TODAY'S ATTENDANCE WITH SCHEDULE
# ====================================================

@staff_bp.route("/attendance", methods=["GET"])
@staff_bp.route("/attendance/today", methods=["GET"])
@check_authorization("OWNER", "FINANCE", "HRD", "CASHIER")
def get_attendance(user):
    """Get attendance data for a selected date (includes scheduled staff not checked in yet)"""
    try:
        date_param = request.args.get("date")
        selected_date = (
            datetime.fromisoformat(date_param).date()
            if date_param
            else waktu_wib().date()
        )
        
        # Query: Staff Schedule + Staff + Shift + LEFT JOIN Attendance
        query = db.session.query(
            Staff.id,
            Staff.employee_code,
            Staff.full_name,
            Staff.department,
            Staff.position,
            Shift.shift_name,
            Shift.start_time,
            Shift.end_time,
            Shift.tolerance_minutes,
            Attendance.clock_in,
            Attendance.clock_out,
            Attendance.status,
            Attendance.late_minutes,
            Attendance.work_minutes,
            StaffSchedule.id.label("schedule_id"),
            StaffSchedule.shift_id,
            StaffSchedule.schedule_date
        ).join(
            StaffSchedule, Staff.id == StaffSchedule.staff_id
        ).join(
            Shift, StaffSchedule.shift_id == Shift.id
        ).outerjoin(
            Attendance, StaffSchedule.id == Attendance.schedule_id
        ).filter(
            StaffSchedule.schedule_date == selected_date,
            Staff.status == "ACTIVE"
        ).order_by(
            Shift.start_time.asc(),
            Staff.full_name.asc(),
            StaffSchedule.id.asc()
        ).all()
        
        data = []
        for row in query:
            work_minutes = calculate_work_minutes(
                row.clock_in,
                row.clock_out,
                row.work_minutes or 0
            )
            data.append({
                "id": row.id,
                "schedule_id": row.schedule_id,
                "shift_id": row.shift_id,
                "schedule_date": row.schedule_date.isoformat(),
                "employee_code": row.employee_code,
                "full_name": row.full_name,
                "department": row.department,
                "position": row.position,
                "shift_name": row.shift_name,
                "start_time": row.start_time.isoformat(),
                "end_time": row.end_time.isoformat(),
                "tolerance_minutes": row.tolerance_minutes,
                "clock_in": row.clock_in.isoformat() if row.clock_in else None,
                "clock_out": row.clock_out.isoformat() if row.clock_out else None,
                "status": row.status or "NOT_CHECKED_IN",
                "late_minutes": row.late_minutes or 0,
                "work_minutes": work_minutes
            })
        
        return jsonify({
            "success": True,
            "data": data,
            "schedule_date": selected_date.isoformat(),
            "total_scheduled": len(data)
        }), 200
    except ValueError:
        return jsonify({
            "success": False,
            "message": "Tanggal jadwal tidak valid"
        }), 400
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500
        
@staff_bp.route("/attendance/history", methods=["GET"])
@check_authorization("OWNER", "FINANCE", "HRD")
def attendance_history(user):

    try:

        rows = db.session.query(

            Staff.full_name,
            Attendance.attendance_date,
            Attendance.clock_in,
            Attendance.clock_out,
            Attendance.status,
            Shift.shift_name

        ).join(
            StaffSchedule,
            Attendance.schedule_id == StaffSchedule.id
        ).join(
            Staff,
            Staff.id == Attendance.staff_id
        ).join(
            Shift,
            Shift.id == StaffSchedule.shift_id
        ).order_by(
            Attendance.attendance_date.desc(),
            Staff.full_name.asc()
        ).all()

        data = []

        for row in rows:

            data.append({

                "name": row.full_name,
                "date": row.attendance_date.isoformat(),
                "clock_in": row.clock_in.isoformat() if row.clock_in else None,
                "clock_out": row.clock_out.isoformat() if row.clock_out else None,
                "shift": row.shift_name,
                "status": row.status

            })

        return jsonify({
            "success": True,
            "data": data
        })

    except Exception as e:

        return jsonify({
            "success": False,
            "message": str(e)
        }),500

@staff_bp.route("/attendance/<int:schedule_id>/clock-in", methods=["PATCH"])
@check_authorization("FINANCE")
def clock_in(user, schedule_id):
    """Clock in for a scheduled staff member"""
    try:
        now = waktu_wib()
        today = now.date()

        schedule = StaffSchedule.query.get(schedule_id)
        if not schedule or schedule.schedule_date != today:
            return jsonify({
                "success": False,
                "message": "Jadwal hari ini tidak ditemukan"
            }), 404

        attendance, error = clock_in_schedule(schedule, now)
        if error:
            message, status_code = error
            return jsonify({
                "success": False,
                "message": message
            }), status_code

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Clock in berhasil",
            "data": {
                "status": attendance.status,
                "clock_in": attendance.clock_in.isoformat(),
                "late_minutes": attendance.late_minutes
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/attendance/<int:schedule_id>/clock-out", methods=["PATCH"])
@check_authorization("FINANCE")
def clock_out(user, schedule_id):
    """Clock out for a scheduled staff member"""
    try:
        now = waktu_wib()
        today = now.date()

        schedule = StaffSchedule.query.get(schedule_id)
        if not schedule or schedule.schedule_date != today:
            return jsonify({
                "success": False,
                "message": "Jadwal hari ini tidak ditemukan"
            }), 404

        attendance, error = clock_out_schedule(schedule, now)
        if error:
            message, status_code = error
            return jsonify({
                "success": False,
                "message": message
            }), status_code

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Clock out berhasil",
            "data": {
                "status": attendance.status,
                "clock_out": attendance.clock_out.isoformat(),
                "late_minutes": attendance.late_minutes,
                "work_minutes": attendance.work_minutes
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


def get_current_user_today_schedule(user):
    staff = Staff.query.filter_by(user_id=user.id, status="ACTIVE").first()
    if not staff:
        return None, "Profil Staff untuk akun ini tidak ditemukan"

    schedule = StaffSchedule.query.filter_by(
        staff_id=staff.id,
        schedule_date=waktu_wib().date()
    ).first()
    if not schedule:
        return None, "Jadwal hari ini tidak ditemukan"

    return schedule, None


def serialize_current_attendance(schedule):
    attendance = Attendance.query.filter_by(schedule_id=schedule.id).first()
    if attendance:
        sync_attendance_with_schedule(schedule, attendance)

    now = waktu_wib()
    shift_start, shift_end = schedule_datetime_range(schedule)
    work_minutes = calculate_work_minutes(
        attendance.clock_in if attendance else None,
        attendance.clock_out if attendance else None,
        attendance.work_minutes if attendance else 0
    )

    return {
        "schedule_id": schedule.id,
        "staff_name": schedule.staff.full_name,
        "position": role_label(schedule.staff.user.role.role_name if schedule.staff.user and schedule.staff.user.role else schedule.staff.position),
        "department": role_group(schedule.staff.user.role.role_name if schedule.staff.user and schedule.staff.user.role else schedule.staff.department),
        "schedule_date": schedule.schedule_date.isoformat(),
        "shift_name": schedule.shift.shift_name,
        "start_time": schedule.shift.start_time.strftime("%H:%M"),
        "end_time": schedule.shift.end_time.strftime("%H:%M"),
        "status": attendance.status if attendance else "NOT_CHECKED_IN",
        "clock_in": attendance.clock_in.isoformat() if attendance and attendance.clock_in else None,
        "clock_out": attendance.clock_out.isoformat() if attendance and attendance.clock_out else None,
        "work_minutes": work_minutes,
        "can_clock_in": bool(attendance is None or not attendance.clock_in) and shift_start <= now <= shift_end,
        "can_clock_out": bool(attendance and attendance.clock_in and not attendance.clock_out) and now >= shift_end,
        "clock_in_locked_message": (
            f"Absen masuk dibuka pukul {schedule.shift.start_time.strftime('%H:%M')}"
            if now < shift_start else ""
        ),
        "clock_out_locked_message": (
            f"Absen pulang dibuka pukul {schedule.shift.end_time.strftime('%H:%M')}"
            if now < shift_end else ""
        )
    }


@staff_bp.route("/attendance/me", methods=["GET"])
@check_authorization("EMPLOYEE")
def current_user_attendance(user):
    try:
        schedule, message = get_current_user_today_schedule(user)
        if not schedule:
            return jsonify({
                "success": False,
                "message": message
            }), 404

        data = serialize_current_attendance(schedule)
        db.session.commit()
        return jsonify({
            "success": True,
            "data": data
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/attendance/clock-in", methods=["PATCH"])
@check_authorization("EMPLOYEE")
def clock_in_current_user(user):
    """Clock in for the current logged-in staff profile."""
    try:
        now = waktu_wib()
        schedule, message = get_current_user_today_schedule(user)
        if not schedule:
            return jsonify({
                "success": False,
                "message": message
            }), 404

        attendance, error = clock_in_schedule(schedule, now)
        if error:
            message, status_code = error
            return jsonify({
                "success": False,
                "message": message
            }), status_code

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Clock in berhasil",
            "data": {
                "schedule_id": schedule.id,
                "status": attendance.status,
                "clock_in": attendance.clock_in.isoformat(),
                "late_minutes": attendance.late_minutes,
                **serialize_current_attendance(schedule)
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/attendance/clock-out", methods=["PATCH"])
@check_authorization("EMPLOYEE")
def clock_out_current_user(user):
    """Clock out for the current logged-in staff profile."""
    try:
        now = waktu_wib()
        schedule, message = get_current_user_today_schedule(user)
        if not schedule:
            return jsonify({
                "success": False,
                "message": message
            }), 404

        attendance, error = clock_out_schedule(schedule, now)
        if error:
            message, status_code = error
            return jsonify({
                "success": False,
                "message": message
            }), status_code

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Clock out berhasil",
            "data": {
                "schedule_id": schedule.id,
                "status": attendance.status,
                "clock_out": attendance.clock_out.isoformat(),
                "late_minutes": attendance.late_minutes,
                "work_minutes": attendance.work_minutes,
                **serialize_current_attendance(schedule)
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

# ====================================================
# STATISTICS
# ====================================================

@staff_bp.route("/statistics/today", methods=["GET"])
@check_authorization("OWNER", "FINANCE", "HRD")
def get_today_statistics(user):
    """Get today's statistics"""
    try:
        today = waktu_wib().date()
        
        total_staff = db.session.query(func.count(Staff.id)).filter(
            Staff.status == "ACTIVE"
        ).scalar() or 0

        # Total scheduled active staff for today
        total_scheduled = db.session.query(func.count(StaffSchedule.id)).join(
            Staff, StaffSchedule.staff_id == Staff.id
        ).filter(
            StaffSchedule.schedule_date == today,
            Staff.status == "ACTIVE"
        ).scalar() or 0
        
        # Present count (PRESENT, LATE, COMPLETED)
        present_count = db.session.query(func.count(func.distinct(Attendance.schedule_id))).join(
            StaffSchedule, Attendance.schedule_id == StaffSchedule.id
        ).join(
            Staff, StaffSchedule.staff_id == Staff.id
        ).filter(
            Attendance.attendance_date == today,
            Staff.status == "ACTIVE",
            Attendance.status.in_(["PRESENT", "LATE", "COMPLETED"])
        ).scalar() or 0
        
        # Late count
        late_count = db.session.query(func.count(func.distinct(Attendance.schedule_id))).join(
            StaffSchedule, Attendance.schedule_id == StaffSchedule.id
        ).join(
            Staff, StaffSchedule.staff_id == Staff.id
        ).filter(
            Attendance.attendance_date == today,
            Staff.status == "ACTIVE",
            Attendance.late_minutes > 0
        ).scalar() or 0
        
        # Average late minutes
        avg_late = db.session.query(func.avg(Attendance.late_minutes)).join(
            StaffSchedule, Attendance.schedule_id == StaffSchedule.id
        ).join(
            Staff, StaffSchedule.staff_id == Staff.id
        ).filter(
            Attendance.attendance_date == today,
            Staff.status == "ACTIVE",
            Attendance.late_minutes > 0
        ).scalar() or 0
        
        # Leave/Sick count
        leave_count = db.session.query(func.count(func.distinct(Attendance.schedule_id))).join(
            StaffSchedule, Attendance.schedule_id == StaffSchedule.id
        ).join(
            Staff, StaffSchedule.staff_id == Staff.id
        ).filter(
            Attendance.attendance_date == today,
            Staff.status == "ACTIVE",
            Attendance.status.in_(["LEAVE", "SICK"])
        ).scalar() or 0
        
        return jsonify({
            "success": True,
            "data": {
                "total_staff": total_staff,
                "total_scheduled": total_scheduled,
                "present_count": present_count,
                "attendance_rate": round((present_count / total_scheduled) * 100, 1) if total_scheduled else 0,
                "late_count": late_count,
                "avg_late_minutes": round(avg_late, 0) if avg_late else 0,
                "leave_count": leave_count
            }
        }), 200
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@staff_bp.route("/statistics/month", methods=["GET"])
@check_authorization("OWNER", "FINANCE", "HRD")
def get_month_statistics(user):
    """Get this month's statistics"""
    try:
        today = waktu_wib().date()
        month_start = date(today.year, today.month, 1)
        
        # Next month's first day
        if today.month == 12:
            month_end = date(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
        
        # Total scheduled in month
        total_scheduled = db.session.query(func.count(StaffSchedule.id)).join(
            Staff, StaffSchedule.staff_id == Staff.id
        ).filter(
            StaffSchedule.schedule_date >= month_start,
            StaffSchedule.schedule_date <= month_end,
            Staff.status == "ACTIVE"
        ).scalar() or 0
        
        # Total attended (PRESENT, LATE, COMPLETED)
        total_attended = db.session.query(func.count(func.distinct(Attendance.schedule_id))).join(
            StaffSchedule, Attendance.schedule_id == StaffSchedule.id
        ).join(
            Staff, StaffSchedule.staff_id == Staff.id
        ).filter(
            Attendance.attendance_date >= month_start,
            Attendance.attendance_date <= month_end,
            Staff.status == "ACTIVE",
            Attendance.status.in_(["PRESENT", "LATE", "COMPLETED"])
        ).scalar() or 0
        
        # Total leave/sick (LEAVE, SICK)
        total_leave = db.session.query(func.count(func.distinct(Attendance.schedule_id))).join(
            StaffSchedule, Attendance.schedule_id == StaffSchedule.id
        ).join(
            Staff, StaffSchedule.staff_id == Staff.id
        ).filter(
            Attendance.attendance_date >= month_start,
            Attendance.attendance_date <= month_end,
            Staff.status == "ACTIVE",
            Attendance.status.in_(["LEAVE", "SICK"])
        ).scalar() or 0
        
        # Calculate rate
        attendance_rate = 0
        if total_scheduled > 0:
            attendance_rate = round((total_attended / total_scheduled) * 100, 1)
        
        return jsonify({
            "success": True,
            "data": {
                "total_attended": total_attended,
                "total_leave": total_leave,
                "attendance_rate": attendance_rate
            }
        }), 200
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

# ====================================================
# LEAVE REQUEST
# ====================================================

@staff_bp.route("/leave-request", methods=["GET"])
@check_authorization("OWNER")
def get_leave_requests(user):
    """Get pending leave requests"""
    try:
        requests = LeaveRequest.query.filter_by(status="PENDING").order_by(
            LeaveRequest.created_at.desc()
        ).all()
        data = [serialize_leave_request(r) for r in requests]
        
        return jsonify({
            "success": True,
            "data": data
        }), 200
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/leave-request/manage", methods=["GET"])
@check_authorization("FINANCE", "OWNER")
def manage_leave_requests(user):
    """Get leave request documents for finance management."""
    try:
        status = normalize_cell(request.args.get("status")).upper()
        query = LeaveRequest.query

        if status in {"PENDING", "APPROVED", "REJECTED"}:
            query = query.filter(LeaveRequest.status == status)

        requests = query.order_by(LeaveRequest.created_at.desc()).limit(200).all()

        return jsonify({
            "success": True,
            "data": [serialize_leave_request(r) for r in requests]
        }), 200
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/leave-request", methods=["POST"])
@check_authorization("FINANCE")
def create_leave_request(user):
    """Create leave/sick request for a staff member"""
    try:
        is_multipart = request.content_type and request.content_type.startswith("multipart/form-data")
        data = request.form if is_multipart else (request.get_json(silent=True) or {})
        document_file = request.files.get("document") if is_multipart else None

        if not data.get("staff_id") or not data.get("leave_type"):
            return jsonify({
                "success": False,
                "message": "Staff dan jenis izin wajib dipilih"
            }), 400

        leave_type = normalize_cell(data.get("leave_type")).upper()
        if leave_type not in {"SICK", "LEAVE", "PERMISSION"}:
            return jsonify({
                "success": False,
                "message": "Jenis izin tidak valid"
            }), 400

        staff = Staff.query.get(int(data.get("staff_id")))
        if not staff or staff.status != "ACTIVE":
            return jsonify({
                "success": False,
                "message": "Staff tidak ditemukan atau tidak aktif"
            }), 404

        start_date_value = data.get("start_date") or data.get("schedule_date")
        end_date_value = data.get("end_date") or start_date_value
        if not start_date_value:
            return jsonify({
                "success": False,
                "message": "Tanggal mulai wajib diisi"
            }), 400

        start_date = datetime.fromisoformat(start_date_value).date()
        end_date = datetime.fromisoformat(end_date_value).date()
        if end_date < start_date:
            return jsonify({
                "success": False,
                "message": "Tanggal selesai tidak boleh sebelum tanggal mulai"
            }), 400

        if (end_date - start_date).days > 366:
            return jsonify({
                "success": False,
                "message": "Rentang izin maksimal 1 tahun"
            }), 400

        document_url = save_leave_document(document_file) if document_file else None
        auto_approve = False
        leave_req = LeaveRequest(
            staff_id=staff.id,
            leave_type=leave_type,
            start_date=start_date,
            end_date=end_date,
            reason=normalize_cell(data.get("reason")) or None,
            document_url=document_url or normalize_cell(data.get("document_url")) or None,
            status="APPROVED" if auto_approve else "PENDING",
            reviewed_by=user.id if auto_approve else None,
            reviewed_at=waktu_wib() if auto_approve else None
        )
        db.session.add(leave_req)
        db.session.flush()

        applied_count = 0
        if auto_approve:
            applied_count = apply_leave_request_to_attendance(leave_req)

        db.session.commit()

        if auto_approve:
            message = "Izin/Sakit berhasil disetujui"
            if applied_count == 0:
                message += ", tetapi belum ada jadwal pada rentang tanggal tersebut"
        else:
            message = "Pengajuan izin/sakit berhasil dibuat"

        return jsonify({
            "success": True,
            "message": message,
            "data": {
                "id": leave_req.id,
                "staff_id": leave_req.staff_id,
                "leave_type": leave_req.leave_type,
                "start_date": leave_req.start_date.isoformat(),
                "end_date": leave_req.end_date.isoformat(),
                "document_url": leave_req.document_url,
                "status": leave_req.status,
                "applied_count": applied_count
            }
        }), 201
    except ValueError as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e) or "Data izin/sakit tidak valid"
        }), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/leave-request/<int:request_id>", methods=["PATCH", "PUT"])
@check_authorization("FINANCE")
def update_leave_request(user, request_id):
    """Update a finance-created leave request and optionally replace its document."""
    try:
        leave_req = LeaveRequest.query.get(request_id)
        if not leave_req:
            return jsonify({
                "success": False,
                "message": "Surat izin tidak ditemukan"
            }), 404

        if leave_req.status != "PENDING":
            return jsonify({
                "success": False,
                "message": "Surat izin yang sudah disetujui atau ditolak tidak dapat diedit"
            }), 409

        is_multipart = request.content_type and request.content_type.startswith("multipart/form-data")
        data = request.form if is_multipart else (request.get_json(silent=True) or {})
        document_file = request.files.get("document") if is_multipart else None

        leave_type = normalize_cell(data.get("leave_type") or leave_req.leave_type).upper()
        if leave_type not in {"SICK", "LEAVE", "PERMISSION"}:
            return jsonify({
                "success": False,
                "message": "Jenis izin tidak valid"
            }), 400

        start_date_value = data.get("start_date") or leave_req.start_date.isoformat()
        end_date_value = data.get("end_date") or start_date_value
        start_date = datetime.fromisoformat(start_date_value).date()
        end_date = datetime.fromisoformat(end_date_value).date()

        if end_date < start_date:
            return jsonify({
                "success": False,
                "message": "Tanggal selesai tidak boleh sebelum tanggal mulai"
            }), 400

        if (end_date - start_date).days > 366:
            return jsonify({
                "success": False,
                "message": "Rentang izin maksimal 1 tahun"
            }), 400

        document_url = save_leave_document(document_file) if document_file else None

        leave_req.leave_type = leave_type
        leave_req.start_date = start_date
        leave_req.end_date = end_date
        leave_req.reason = normalize_cell(data.get("reason")) or None
        if document_url:
            leave_req.document_url = document_url

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Surat izin berhasil diperbarui",
            "data": serialize_leave_request(leave_req)
        }), 200
    except ValueError as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e) or "Data surat izin tidak valid"
        }), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@staff_bp.route("/leave-request/<int:request_id>/approve", methods=["PATCH"])
@check_authorization("OWNER")
def approve_leave(user, request_id):
    """Approve leave request"""
    try:
        leave_req = LeaveRequest.query.get(request_id)
        if not leave_req:
            return jsonify({
                "success": False,
                "message": "Permintaan tidak ditemukan"
            }), 404

        if leave_req.status != "PENDING":
            return jsonify({
                "success": False,
                "message": "Permintaan sudah direview"
            }), 409
        
        leave_req.status = "APPROVED"
        leave_req.reviewed_by = user.id
        leave_req.reviewed_at = waktu_wib()

        applied_count = apply_leave_request_to_attendance(leave_req)
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "Permintaan berhasil disetujui",
            "data": {
                "applied_count": applied_count
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@staff_bp.route("/leave-request/<int:request_id>/reject", methods=["PATCH"])
@check_authorization("OWNER")
def reject_leave(user, request_id):
    """Reject leave request"""
    try:
        leave_req = LeaveRequest.query.get(request_id)
        if not leave_req:
            return jsonify({
                "success": False,
                "message": "Permintaan tidak ditemukan"
            }), 404

        if leave_req.status != "PENDING":
            return jsonify({
                "success": False,
                "message": "Permintaan sudah direview"
            }), 409
        
        leave_req.status = "REJECTED"
        leave_req.reviewed_by = user.id
        leave_req.reviewed_at = waktu_wib()
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "Permintaan berhasil ditolak"
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@staff_bp.route("/accounts", methods=["GET"])
def get_staff_accounts():

    result = db.session.execute(text("""
        SELECT
            u.id,
            s.full_name,
            u.username,
            r.role_name,
            CASE
                WHEN u.is_active = 1 THEN 'ACTIVE'
                ELSE 'INACTIVE'
            END AS status,
            u.updated_at
        FROM users u
        LEFT JOIN staff s
            ON s.user_id = u.id
        LEFT JOIN roles r
            ON r.id = u.role_id
        ORDER BY s.full_name
    """))

    data = []

    for row in result:

        data.append({

            "id": row.id,
            "name": row.full_name or "-",
            "username": row.username,
            "role": role_form_code(row.role_name),
            "role_label": role_label(row.role_name),
            "role_group": role_group(row.role_name),
            "status": row.status,
            "last_login": row.updated_at.strftime("%d/%m/%Y %H:%M")
            if row.updated_at else "-"

        })

    return jsonify({
        "success": True,
        "data": data
    })


@staff_bp.route("/roles", methods=["GET"])
def get_staff_roles():
    roles = db.session.execute(text("""
        SELECT role_name, description
        FROM roles
        WHERE UPPER(role_name) NOT IN ('OWNER', 'KASIR', 'HRD')
        ORDER BY role_name
    """)).fetchall()

    data = []
    for row in roles:
        data.append({
            "code": row.role_name,
            "label": role_label(row.role_name),
            "group": role_group(row.role_name)
        })

    data.sort(key=lambda item: role_sort_key(item["code"]))

    return jsonify({
        "success": True,
        "data": data
    })
